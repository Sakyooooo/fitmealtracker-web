# -*- coding: utf-8 -*-
"""
日本食品標準成分表2020年版（八訂）本表 Excel → 軽量JSON 変換スクリプト。

出典: 文部科学省「日本食品標準成分表2020年版（八訂）」第2章（データ）
      https://www.mext.go.jp/a_menu/syokuhinseibun/mext_01110.html
      （本表 xlsx: 20201225-mxt_kagsei-mext_01110_012.xlsx）

可食部100gあたりの エネルギー(kcal)・たんぱく質・脂質・炭水化物 のみを抽出し、
料理トラッカーの「食品検索」用に最小化したJSONを出力する。

入力 : scripts/_render/food_comp_raw.xlsx  （MEXT本表・combinedシート）
出力 : public/data/food_composition.json

列マッピング（1始まり, 八訂本表 combinedシート）:
  1: 食品群  2: 食品番号  4: 食品名
  7: ENERC_KCAL(kcal)  10: PROT-(たんぱく質)  13: FAT-(脂質)  21: CHOCDF-(炭水化物)
"""
import json
import os
import re
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "scripts", "_render", "food_comp_raw.xlsx")
OUT_DIR = os.path.join(ROOT, "public", "data")
OUT = os.path.join(OUT_DIR, "food_composition.json")

# 列インデックス（0始まり）
COL_GROUP, COL_CODE, COL_NAME = 0, 1, 3
COL_KCAL, COL_PROT, COL_FAT, COL_CARB = 6, 9, 12, 20

CATEGORY = {
    "01": "穀類", "02": "いも及びでん粉類", "03": "砂糖及び甘味類", "04": "豆類",
    "05": "種実類", "06": "野菜類", "07": "果実類", "08": "きのこ類", "09": "藻類",
    "10": "魚介類", "11": "肉類", "12": "卵類", "13": "乳類", "14": "油脂類",
    "15": "菓子類", "16": "し好飲料類", "17": "調味料及び香辛料類", "18": "調理加工食品類",
}


def parse_num(v):
    """成分表の数値表記を float へ。'-'/空=None、'Tr'=0、'(x)'(推定)=x。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "-", "ND"):
        return None
    if s in ("Tr", "(Tr)", "微量"):
        return 0.0
    s = s.strip("()").replace(",", "").replace("*", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def kata_to_hira(s: str) -> str:
    out = []
    for ch in s:
        code = ord(ch)
        if 0x30A1 <= code <= 0x30F6:  # カタカナ → ひらがな
            out.append(chr(code - 0x60))
        else:
            out.append(ch)
    return "".join(out)


def search_key(name: str) -> str:
    """空白・記号を除去し、カタカナをひらがなへ寄せた検索用キー。"""
    s = name.lower()
    s = re.sub(r"[\s　・,，.．。\-\[\]（）()【】「」<>＜＞/／]", "", s)
    return kata_to_hira(s)


def main():
    if not os.path.exists(SRC):
        raise SystemExit(
            f"入力が見つかりません: {SRC}\n"
            "MEXT本表 xlsx を取得して上記パスに置いてください。\n"
            "URL: https://www.mext.go.jp/content/20201225-mxt_kagsei-mext_01110_012.xlsx"
        )

    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]  # combined 本表

    items = []
    skipped = 0
    for row in ws.iter_rows(min_row=13):
        cells = [c.value for c in row]
        if len(cells) <= COL_CARB:
            continue
        code = cells[COL_CODE]
        name = cells[COL_NAME]
        if code is None or name is None:
            continue
        code = str(code).strip()
        if not re.fullmatch(r"\d{5}", code):
            continue
        kcal = parse_num(cells[COL_KCAL])
        if kcal is None:
            skipped += 1
            continue  # エネルギー不明の食品は検索対象にしない
        name = re.sub(r"\s+", " ", str(name).replace("　", " ")).strip()
        group = str(cells[COL_GROUP]).strip().zfill(2) if cells[COL_GROUP] is not None else code[:2]

        items.append({
            "id": code,
            "name": name,
            "category": CATEGORY.get(group, ""),
            "kcal": round(kcal, 1),
            "p": parse_num(cells[COL_PROT]),
            "f": parse_num(cells[COL_FAT]),
            "c": parse_num(cells[COL_CARB]),
            "searchKey": search_key(name),
        })

    payload = {
        "meta": {
            "source": "文部科学省 日本食品標準成分表2020年版（八訂）第2章（データ）",
            "sourceUrl": "https://www.mext.go.jp/a_menu/syokuhinseibun/mext_01110.html",
            "unit": "可食部100gあたり",
            "note": "エネルギーはkcal、たんぱく質(p)/脂質(f)/炭水化物(c)はg。炭水化物は差引き法(CHOCDF)。",
            "count": len(items),
        },
        "items": items,
    }

    os.makedirs(OUT_DIR, exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))

    size_kb = os.path.getsize(OUT) // 1024
    print(f"OK: {len(items)} 品 を書き出し（kcal無しでスキップ {skipped}）-> {OUT} ({size_kb}KB)")
    # サンプル確認
    for kw in ("精白米", "鶏卵", "若どり"):
        for it in items:
            if kw in it["name"]:
                print(f"  例 {it['id']} {it['name']}: {it['kcal']}kcal P{it['p']}/F{it['f']}/C{it['c']}")
                break


if __name__ == "__main__":
    main()
